import os
import shutil
import signal
import subprocess
import sys
from random import randint
from statistics import mean
from time import sleep

from client import Client
from src.config import INIT_PORT, HOP_SERVER_PORT, HYPERCUBE_SIZE
from src.utils import reset_hops, get_hops, NODES, log
from openpyxl import load_workbook

IPFS_CLIENT_ADDRESS = '/ip4/127.0.0.1/tcp/5001'

OBJECTS = [10, 100, 1000]
PINSEARCH_ATTEMPS = 50
SUPERSET_ATTEMPS = 50
SUPERSET_LIMIT = 10

TEST_FILES = './test_files/'
TEMPLATE = 'test_template.xlsx'
RESULTS_FILE = 'results/test_results.xlsx'
SHEET_PIN = 'pin'
SHEET_SUPERSET = 'superset'
FIRST_COLUMN = 2

document = load_workbook(RESULTS_FILE)
pin_sheet = document[SHEET_PIN]
superset_sheet = document[SHEET_SUPERSET]


for col in range(0, len(OBJECTS)):
    servers = [subprocess.Popen([sys.executable, 'hops_counter.py'])]
    log('HOPS SERVER', 'STARTING', 'PORT {}'.format(str(HOP_SERVER_PORT)))

    for pid in range(0, NODES):
        servers.append(subprocess.Popen([sys.executable, 'server.py', str(INIT_PORT + pid)]))
        log('SERVER {}'.format(str(pid)), 'STARTED', 'PORT {}'.format(str(INIT_PORT + pid)))

    sleep(40)

    # clears the test directory
    if os.path.exists(TEST_FILES):
        shutil.rmtree(TEST_FILES)
    os.makedirs(TEST_FILES)

    # creates random files and adds them from a random nodes
    for i in range(0, OBJECTS[col]):
        client = Client(IPFS_CLIENT_ADDRESS, randint(0, NODES-1))
        with open(TEST_FILES + str(i), 'wb') as random_file:
            random_file.write(os.urandom(512))
        client.add_obj(TEST_FILES + str(i), randint(0, NODES-1))
        client.close()

    # pinsearch test
    pinsearch_hops = []
    for i in range(0, PINSEARCH_ATTEMPS):
        client = Client(IPFS_CLIENT_ADDRESS, randint(0, NODES-1))
        reset_hops()
        client.pin_search(randint(0, NODES-1))
        pinsearch_hops.append(get_hops())
        client.close()

    # superset test
    superset_hops = []
    for i in range(0, SUPERSET_ATTEMPS):
        client = Client(IPFS_CLIENT_ADDRESS, randint(0, NODES-1))
        reset_hops()
        client.superset_search(randint(0, NODES-1), SUPERSET_LIMIT)
        superset_hops.append(get_hops())
        client.close()

    pin_sheet.cell(row=HYPERCUBE_SIZE, column=col+FIRST_COLUMN).value = mean(pinsearch_hops)
    superset_sheet.cell(row=HYPERCUBE_SIZE, column=col+FIRST_COLUMN).value = mean(superset_hops)

    sleep(5)

    for pid in servers:
        os.kill(pid.pid, signal.SIGTERM)

document.save(RESULTS_FILE)
